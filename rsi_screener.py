import pandas as pd
import yfinance as yf
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, time, requests

# ─── CONFIG ───────────────────────────────────────────────────────────────────
RSI_PERIOD      = 14
MONTHLY_RSI_MIN = 60
WEEKLY_RSI_MIN  = 60
DAILY_RSI_MIN   = 40
DAILY_RSI_MAX   = 45
OUTPUT_DIR      = "output"

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID   = os.environ.get("TELEGRAM_CHAT_ID", "")
# ──────────────────────────────────────────────────────────────────────────────


# ══════════════════════════════════════════════════════════════════════════════
#  RSI CALCULATION
# ══════════════════════════════════════════════════════════════════════════════

def compute_rsi(series: pd.Series, period: int = 14) -> float:
    delta    = series.diff()
    gain     = delta.clip(lower=0)
    loss     = -delta.clip(upper=0)
    avg_gain = gain.ewm(com=period - 1, min_periods=period).mean()
    avg_loss = loss.ewm(com=period - 1, min_periods=period).mean()
    rs       = avg_gain / avg_loss.replace(0, np.nan)
    rsi      = 100 - (100 / (1 + rs))
    return round(float(rsi.iloc[-1]), 2) if not rsi.empty else None


def load_symbols(filepath: str) -> list:
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip().str.upper()
    col = next((c for c in df.columns if "SYMBOL" in c), None)
    if col is None:
        raise ValueError(f"No 'Symbol' column in {filepath}")
    return df[col].dropna().astype(str).str.strip().str.upper().tolist()


def fetch_rsi(symbol: str) -> dict:
    ticker = f"{symbol}.NS"
    try:
        raw = yf.download(ticker, period="5y", interval="1d",
                          auto_adjust=True, progress=False)
        if raw.empty or len(raw) < RSI_PERIOD + 5:
            return None
        close_daily   = raw["Close"].squeeze()
        daily_rsi     = compute_rsi(close_daily)
        weekly_close  = close_daily.resample("W").last().dropna()
        weekly_rsi    = compute_rsi(weekly_close) if len(weekly_close) >= RSI_PERIOD + 5 else None
        monthly_close = close_daily.resample("ME").last().dropna()
        monthly_rsi   = compute_rsi(monthly_close) if len(monthly_close) >= RSI_PERIOD + 5 else None
        return {
            "Symbol":        symbol,
            "Current Price": round(float(close_daily.iloc[-1]), 2),
            "Daily RSI":     daily_rsi,
            "Weekly RSI":    weekly_rsi,
            "Monthly RSI":   monthly_rsi,
        }
    except Exception as e:
        print(f"  WARNING {symbol}: {e}")
        return None


def matches_filter(row: dict) -> bool:
    d, w, m = row.get("Daily RSI"), row.get("Weekly RSI"), row.get("Monthly RSI")
    if any(v is None for v in [d, w, m]):
        return False
    return m > MONTHLY_RSI_MIN and w > WEEKLY_RSI_MIN and DAILY_RSI_MIN < d < DAILY_RSI_MAX


def tv_link(symbol: str) -> str:
    return f"https://www.tradingview.com/chart/?symbol=NSE:{symbol}"


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════

def _border(color="D0D0D0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _border("FFFFFF")

def style_cell(cell, align="center"):
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _border()

def highlight_rsi(cell, value, low, high):
    if value > high:
        cell.fill = PatternFill("solid", start_color="C6EFCE")
    elif value >= low:
        cell.fill = PatternFill("solid", start_color="FFEB9C")

def build_sheet(ws, results: list, label: str):
    run_date = datetime.now().strftime("%d-%b-%Y %H:%M IST")
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = f"{label}  |  RSI Screener  |  Run: {run_date}  |  {len(results)} stocks matched"
    t.font      = Font(bold=True, name="Arial", size=11, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color="243F60")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers    = ["#", "Symbol", "TradingView Link", "Current Price (₹)", "Daily RSI", "Weekly RSI", "Monthly RSI"]
    col_widths = [5, 14, 22, 20, 12, 12, 13]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        style_header(c)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    for rn, stock in enumerate(results, 3):
        ws.row_dimensions[rn].height = 18
        sym      = stock["Symbol"]
        row_data = [rn-2, sym, tv_link(sym), stock["Current Price"],
                    stock["Daily RSI"], stock["Weekly RSI"], stock["Monthly RSI"]]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=rn, column=ci, value=val)
            style_cell(c, align="left" if ci in (2, 3) else "center")
            if ci == 4:
                c.number_format = '#,##0.00'
            elif ci == 5:
                highlight_rsi(c, val, DAILY_RSI_MIN, DAILY_RSI_MAX)
            elif ci == 6:
                highlight_rsi(c, val, WEEKLY_RSI_MIN, 80)
            elif ci == 7:
                highlight_rsi(c, val, MONTHLY_RSI_MIN, 80)

    ws.freeze_panes = "A3"
    lr = len(results) + 4
    ws.cell(row=lr, column=1, value="Legend:").font = Font(bold=True, name="Arial", size=9)
    for ci, (txt, col) in enumerate([("Monthly RSI>60","C6EFCE"),("Weekly RSI>60","C6EFCE"),("Daily RSI 40-45","FFEB9C")], 2):
        c = ws.cell(row=lr, column=ci, value=txt)
        c.fill      = PatternFill("solid", start_color=col)
        c.font      = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center")


# ══════════════════════════════════════════════════════════════════════════════
#  TELEGRAM ALERT
# ══════════════════════════════════════════════════════════════════════════════

def _tg_send(text: str):
    """Send one message chunk to Telegram."""
    url  = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    resp = requests.post(url, json={
        "chat_id":    TELEGRAM_CHAT_ID,
        "text":       text,
        "parse_mode": "Markdown",
    }, timeout=15)
    if resp.status_code != 200:
        print(f"  Telegram error: {resp.status_code} — {resp.text}")


def send_telegram(all_matched: list, run_date: str):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("  Telegram: credentials not set — skipping")
        return

    total = len(all_matched)

    if total == 0:
        msg = (
            f"📊 *NSE RSI Screener — {run_date}*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"❌ *No stocks matched* the RSI filter today\.\n\n"
            f"_Conditions: Monthly RSI \>60 \| Weekly RSI \>60 \| Daily RSI 40–45_"
        )
        _tg_send(msg)
        print("  Telegram: sent (no matches) ✓")
        return

    # ── Header message ─────────────────────────────────────────────────────────
    header = (
        f"📊 *NSE RSI Screener — {run_date}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ *{total} stock(s) matched* all RSI conditions\n"
        f"_NSE500 \+ MicroCap250_\n\n"
        f"*Conditions:*\n"
        f"  • Monthly RSI \> 60\n"
        f"  • Weekly RSI \> 60\n"
        f"  • Daily RSI \> 40 and \< 45\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    _tg_send(header)

    # ── One message per stock ──────────────────────────────────────────────────
    for i, s in enumerate(all_matched, 1):
        sym   = s["Symbol"]
        link  = tv_link(sym)
        price = f"₹{s['Current Price']:,.2f}"
        msg   = (
            f"*{i}\. {sym}*\n"
            f"💰 Price      : `{price}`\n"
            f"📅 Daily RSI  : `{s['Daily RSI']:.2f}`\n"
            f"📆 Weekly RSI : `{s['Weekly RSI']:.2f}`\n"
            f"🗓 Monthly RSI: `{s['Monthly RSI']:.2f}`\n"
            f"📈 [Open Chart]({link})"
        )
        _tg_send(msg)
        time.sleep(0.4)   # avoid Telegram rate limit

    # ── Footer ─────────────────────────────────────────────────────────────────
    footer = (
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"_Excel report committed to GitHub repo\._\n"
        f"_Generated automatically every Saturday 09:05 AM IST_"
    )
    _tg_send(footer)
    print(f"  Telegram: {total + 2} messages sent ✓")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    run_date = datetime.now().strftime("%d-%b-%Y")

    sources = {
        "NSE500":      "data/nse500list.xlsx",
        "MicroCap250": "data/Micro250list.xlsx",
    }

    wb          = Workbook()
    all_matched = []
    sheet_index = 0

    for sheet_name, filepath in sources.items():
        if not os.path.exists(filepath):
            print(f"WARNING: {filepath} not found — skipping")
            continue

        print(f"\n{'='*55}")
        print(f"Loading: {filepath}")
        symbols = load_symbols(filepath)
        seen    = set()
        symbols = [s for s in symbols if not (s in seen or seen.add(s))]
        print(f"  {len(symbols)} unique symbols")

        matched = []
        for i, sym in enumerate(symbols, 1):
            print(f"  [{i:>3}/{len(symbols)}] {sym}", end="  ")
            data = fetch_rsi(sym)
            if data and matches_filter(data):
                matched.append(data)
                print(f"MATCH  D:{data['Daily RSI']}  W:{data['Weekly RSI']}  M:{data['Monthly RSI']}")
            else:
                print("-")
            time.sleep(0.3)

        print(f"\n  {len(matched)} matched from {sheet_name}")
        all_matched.extend(matched)

        ws = wb.active if sheet_index == 0 else wb.create_sheet(title=sheet_name)
        if sheet_index == 0:
            ws.title = sheet_name
        build_sheet(ws, matched, sheet_name)
        sheet_index += 1

    if sheet_index > 1 and all_matched:
        build_sheet(wb.create_sheet(title="ALL_MATCHES"), all_matched, "All Sources Combined")

    date_str    = datetime.now().strftime("%Y-%m-%d")
    report_path = os.path.join(OUTPUT_DIR, f"RSI_Screener_{date_str}.xlsx")
    wb.save(report_path)
    print(f"\nReport saved: {report_path}")

    print("\nSending Telegram alert...")
    send_telegram(all_matched, run_date)
    print("Done.")


if __name__ == "__main__":
    main()
